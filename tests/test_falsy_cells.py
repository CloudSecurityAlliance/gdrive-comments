"""A spreadsheet's `FALSE` is a **boolean**, and it meant the opposite of a typed one.

`openpyxl` returns a TRUE/FALSE cell as Python `True`/`False`, not as text — verified, not
assumed. `_norm` then collapsed it with `str(text or "")`, and `False or ""` is `""`, so a
boolean `FALSE` read as **blank**: no change. A *typed* `FALSE` read as `REVERSE`: reopen the
thread.

Same author intent, opposite behaviour, decided by a cell's type — which the person filling in
the register cannot see and did not choose. And the asymmetry is the tell: `True or ""` is
`True`, so `str()` gave `"True"` and the ACT path worked **by accident**. Only the false branch
was broken, which is why nothing noticed.

Two concrete losses, both the silent-guess outcome `_apply.py` exists to prevent:

  * `resolve_comment=FALSE` as a boolean — a deliberate reopen that silently did nothing.
  * `delete_comment=FALSE` as a boolean — silently no-op instead of hitting the deliberate
    "there is no undelete" refusal, so the user never learned their instruction was impossible.

The xlsx register ships a **dropdown offering `FALSE`**, so a boolean is the value a user is
*most* likely to produce. The most-taken path was the broken one.

`0` and `0.0` had it too: `0 or ""` is `""`. A numeric cell is not exotic — a column formatted
as a number, or a CSV round-tripped through a tool that types its columns, produces exactly
that.

Guarded here, plus the companion defect: the **tool description said the opposite of the code**
(see `TestTheDescriptionMatchesTheCode`), which is how a model gets talked into reopening every
resolved thread while following the documentation exactly.
"""
from __future__ import annotations

import asyncio
import re

import openpyxl
import pytest

from csa_google_workspace import Workspace
from csa_google_workspace._apply import ACT, NONE, REVERSE, _norm, decision, truthy
from csa_google_workspace.backend import FakeBackend
from csa_google_workspace.mcp import settings_from_env
from csa_google_workspace.mcp.server import create_server
from csa_google_workspace.policy import PolicyBackend

DOC = "d1"


def build():
    """Six threads, the even-numbered ones resolved — so a reopen has something to act on."""
    cs = {DOC: [{"id": f"t{i}", "content": f"Point {i}", "author": {"displayName": "A"},
                 "createdTime": "2026-08-20T10:00:00Z", "resolved": i % 2 == 0, "replies": []}
                for i in range(6)]}
    backend = FakeBackend(
        {DOC: {"id": DOC, "name": "Draft", "mimeType": "application/vnd.google-apps.document"}},
        documents={DOC: {"body": {"content": []}}}, comments=cs)
    st = settings_from_env({"CSA_GW_ALLOWLIST_READ": "*", "CSA_GW_ALLOWLIST_MODIFY": "*",
                            "CSA_GW_PROFILE": "full"})
    return create_server(lambda: Workspace(PolicyBackend(backend, st.policy)), settings=st), backend


def call(app, name, **args):
    return asyncio.run(app.call_tool(name, {"fileId": DOC, **args})).structured_content


class TestATypeMustNotChangeTheMeaning:
    """The whole defect, at the unit level."""

    @pytest.mark.parametrize("value", [False, 0, 0.0])
    def test_a_falsy_cell_still_means_reverse(self, value):
        assert decision(value) is REVERSE, (
            f"{value!r} read as {decision(value)!r}; a spreadsheet FALSE is an instruction, "
            f"not an absence")

    @pytest.mark.parametrize("value", [True, 1, 1.0])
    def test_a_truthy_cell_still_means_act(self, value):
        assert decision(value) is ACT

    @pytest.mark.parametrize("value", [None, "", "   "])
    def test_only_genuine_absence_means_none(self, value):
        assert decision(value) is NONE

    @pytest.mark.parametrize("typed,native", [("FALSE", False), ("TRUE", True),
                                              ("0", 0), ("1", 1)])
    def test_typed_and_native_agree(self, typed, native):
        """A .csv and an .xlsx of the same register must do the same thing."""
        assert decision(typed) is decision(native)

    def test_norm_does_not_swallow_false(self):
        """The root cause, named: `str(text or "")` turns False into "".

        Kept as its own assertion because the next person to touch `_norm` will reach for
        `or ""` again — it is the obvious way to write it.
        """
        assert _norm(False) != ""
        assert _norm(0) != ""
        assert _norm(None) == "", "a genuine absence must still normalise to empty"


class TestCompletedMarkersAreUnaffected:
    """`truthy` gates the `*_completed` columns and is read as a boolean, so None and False
    behave identically there. Asserted rather than assumed, because widening `_norm` for the
    decision columns changes what `truthy` returns for the same inputs."""

    @pytest.mark.parametrize("value", [False, 0, None, ""])
    def test_a_falsy_marker_is_not_done(self, value):
        assert not truthy(value)

    @pytest.mark.parametrize("value", [True, 1, "yes", "TRUE"])
    def test_a_truthy_marker_is_done(self, value):
        assert truthy(value)


class TestEndToEndThroughAnXlsxRegister:
    """The path a real user takes: export to .xlsx, pick FALSE from the dropdown, hand it back."""

    def _register(self, app, tmp_path, resolve_value):
        out = call(app, "export_comments", destination="xlsx", path=str(tmp_path / "r.xlsx"))
        path = out["written_path"]
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        col = header.index("resolve_comment") + 1
        thread_col = header.index("thread_id") + 1
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=thread_col).value == "t0":     # a RESOLVED thread
                ws.cell(row=row, column=col).value = resolve_value
        wb.save(path)
        return path

    def test_a_boolean_false_reopens_the_thread(self, tmp_path):
        app, backend = build()
        path = self._register(app, tmp_path, False)
        assert backend._comments[(DOC, "t0")]["resolved"] is True, "fixture precondition"
        out = call(app, "apply_comment_actions", path=path, apply=True)
        assert out["reopened"] == 1, f"boolean FALSE did nothing: {out['rows']}"
        assert backend._comments[(DOC, "t0")]["resolved"] is False

    def test_a_typed_false_does_the_same_thing(self, tmp_path):
        """The comparison that makes it a bug rather than a preference."""
        app, backend = build()
        path = self._register(app, tmp_path, "FALSE")
        out = call(app, "apply_comment_actions", path=path, apply=True)
        assert out["reopened"] == 1
        assert backend._comments[(DOC, "t0")]["resolved"] is False

    def test_a_boolean_false_on_delete_is_refused_not_ignored(self, tmp_path):
        """There is no undelete, so FALSE must fail loudly. Silently ignoring it left the user
        believing a restore had happened."""
        app, _ = build()
        out = call(app, "export_comments", destination="xlsx", path=str(tmp_path / "d.xlsx"))
        path = out["written_path"]
        wb = openpyxl.load_workbook(path); ws = wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        ws.cell(row=2, column=header.index("delete_comment") + 1).value = False
        wb.save(path)
        applied = call(app, "apply_comment_actions", path=path, apply=True)
        assert applied["failed"] == 1, f"a boolean FALSE on delete was ignored: {applied['rows']}"


class TestTheDescriptionMatchesTheCode:
    """The companion defect, and the dangerous half.

    The tool description told a model that `false` means "leave it". The code reopens the
    thread — posting a visible action reply, under the user's name, on a thread somebody
    deliberately closed. A model following the documentation exactly, and pre-filling FALSE on
    every row it did not want to touch, reopens every resolved thread in the document.

    `_apply.py` records that exact failure as the reason the three-state type exists. The
    type-level defence was built; the description then invited the same input through the front
    door.
    """

    @pytest.fixture
    def description(self):
        app, _ = build()
        tools = {t.name: (t.description or "") for t in asyncio.run(app.list_tools())}
        return tools["apply_comment_actions"]

    def test_it_does_not_claim_false_leaves_the_thread_alone(self, description):
        assert not re.search(r"(empty or false|false)\s+means?\s+leave", description, re.I), (
            "the description says FALSE leaves the thread alone; the code reopens it")

    def test_it_says_what_false_actually_does(self, description):
        assert re.search(r"reopen", description, re.I), (
            "a caller cannot know FALSE reopens unless the description says so")

    def test_delete_comment_is_documented_at_all(self, description):
        """An irreversible action that strips both text and author, and the description did not
        mention the column existed."""
        assert "delete_comment" in description

    def test_the_column_count_is_right(self, description):
        """It said "two columns are yours to fill in". There are three."""
        assert not re.search(r"two columns are yours", description, re.I)
