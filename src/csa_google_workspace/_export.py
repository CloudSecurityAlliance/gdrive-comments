"""Comment export: flat rows, ready for a spreadsheet or for another tool entirely.

Two audiences, and **neither of them is an AI**, which is the reason this exists rather than
leaving people to compose three tools:

- somebody who would rather work in a **spreadsheet**. A register of every thread, with the text
  each one is about, is how document review has always been done, and this makes that better
  rather than obsolete.
- **another tool** — a notebook, a BI query, `grep`. Flat rows with a thread id feed anything;
  nested JSON does not.

**Flat, one row per comment AND per reply**, with `reply_to` naming the thread. That is the
lossless shape: one-row-per-thread is a trivial group-by away, and the reverse is not recoverable.
`columns` is returned in order so writing a sheet is a loop rather than a judgement call.

The interesting part is the last column, and it is per-type, because *what a comment points at*
is a different fact for each:

    Docs     `quoted_text`  - the passage, from Drive's quotedFileContent
    Sheets   `cell_text`    - what the cell actually HOLDS. A comment on B11 is meaningless in
                              a register unless the register also says B11 is "Q3 revenue".

And the Sheets half produces an unexpected answer to the multi-tab problem (TODO.md D3). We
cannot know which tab a comment is on — the XLSX export carries no correlation from a
`threadedComments` member back to a sheet name. But we can report that cell **on every tab**,
and the *content* then tells a human which it was: if B11 is empty on Summary and reads
"Q3 revenue" on Detail, the comment is about Detail. That resolves the ambiguity in practice
without the code guessing, which beats both guessing and refusing.
"""
from __future__ import annotations

from typing import Any

# What the export REPORTS.
# `tab` sits immediately before `cell` so `tab | cell | cell_text` reads as one thought. It is
# None for a document, for an unanchored comment, and for a spreadsheet whose relationship graph
# could not be walked - those are three different reasons and none of them is "the first tab".
REPORTED = ("thread_id", "reply_to", "author", "created_time", "resolved", "text",
            # `anchor_state` sits beside `anchored` because it is the same fact at higher
            # resolution: the boolean says whether there is a passage, this says which of the
            # four ways it is attached - including `quote_only`, which reads as file-level to
            # anything looking only at the boolean's OLD meaning (#372).
            "quoted_text", "anchored", "anchor_state",
            "tab", "cell", "cell_text", "cell_text_by_tab",
            "row_header", "column_header",
            # Off unless asked for, and dropped from the header by `used_columns` when empty —
            # a register with five blank columns is harder to read than one without them.
            "context", "context_kind", "context_note", "paragraph", "heading_path")

# What somebody FILLS IN, and what the importer ticks. Always empty on export - they are the
# point of the register being a worksheet rather than a printout.
#
# Two markers rather than none, and the reason is the crash: `*_completed` lets a re-run skip
# finished rows without asking Google about every one of them. It is the FAST path, not the
# authority - see `apply_comment_actions`, which also checks the live thread, because the
# interesting failure is posting a reply and dying before the tick is written.
#   reply_comment    text to post. Empty means no reply.
#   resolve_comment  TRUE resolves, FALSE reopens, empty leaves it alone. Three states, and
#                    the blank one has to stay "I did not decide" - otherwise every untouched
#                    row would reopen every resolved thread.
#   delete_comment   TRUE removes it. For spam. The sharpest action here: Drive's soft delete
#                    strips the content AND the author, permanently. `comment.delete` is ON by
#                    default, and on the profile ladder it arrives at `fileOrganizer`.
ACTIONS = ("reply_comment", "resolve_comment", "delete_comment")
COMPLETED = ("reply_comment_completed", "resolve_comment_completed",
             "delete_comment_completed")
COLUMNS = REPORTED + ACTIONS + COMPLETED


def _row(**kw: Any) -> dict[str, Any]:
    """Every row carries every column, so a spreadsheet write never has ragged rows."""
    row: dict[str, Any] = dict.fromkeys(COLUMNS)
    row.update(kw)
    return row


def _cell_lookup(document: Any) -> tuple[dict[str, list[list[Any]]], list[str]]:
    """({tab: grid}, tabs) for a spreadsheet; ({}, []) for anything else.

    One read per tab, done once for the whole export rather than once per comment - a register
    of forty comments must not be forty API calls.
    """
    # SPREADSHEET tabs, which are plain title strings and are used as dict keys below. The
    # `str` filter is a guard, not tidiness: a Doc briefly had a `tabs` property returning
    # dicts, and this line raised `unhashable type: 'dict'` two calls later with nothing
    # pointing back here. Filtering keeps a same-named property on another type from crashing
    # an export, and `Doc.document_tabs` is now named so the collision cannot recur.
    tabs = [tab for tab in (getattr(document, "tabs", []) or []) if isinstance(tab, str)]
    if not tabs:
        return {}, []
    grids: dict[str, list[list[Any]]] = {}
    for tab in tabs:
        try:
            grids[tab] = document.values(tab) or []
        except Exception:                       # noqa: BLE001 - a tab we cannot read is not
            grids[tab] = []                     # a reason to fail the whole export
    return grids, tabs


def _context_cells(contexts: list | None, i: int) -> dict[str, Any]:
    """The context columns for row `i`, or empty ones.

    Flattened to text because a register is a spreadsheet: `heading_path` joins with ` > ` and
    `paragraph` reads "6 of 9", which is what a person scanning a column wants. The structured
    form stays on the MCP surface, where a consumer can branch on it.
    """
    ctx = contexts[i] if contexts and i < len(contexts) else None
    if ctx is None:
        return {}
    return {"context": flatten(ctx.text), "context_kind": ctx.kind,
            "context_note": ctx.note,
            "paragraph": (f"{ctx.paragraph_index} of {ctx.paragraph_total}"
                          if ctx.paragraph_index else ""),
            "heading_path": " > ".join(ctx.heading_path)}


def _note_count(document: Any) -> int:
    """How many cell notes this file has, or 0 — never an exception.

    Best-effort on purpose. This runs while building a register, and a notes lookup that failed
    must not lose the comments: the caveat is a courtesy, the register is the deliverable. A
    document type with no `notes` (a Doc, a deck) simply has none.
    """
    notes = getattr(document, "notes", None)
    if notes is None:
        return 0
    try:
        return len(list(notes))
    except Exception:                      # noqa: BLE001 - see the docstring
        return 0


def _headers(grid: list[list[Any]], row: int, col: int) -> tuple[str | None, str | None]:
    """The row and column labels for a cell — `(row_header, column_header)`.

    *"B11"* is useless in a register and *"B11, which reads 388000"* is better, but the thing
    that actually makes a cell comment interpretable is *"in the row labelled Southwest, column
    Q3 actual"* — and a comment on the WRONG cell is only detectable against its neighbours.
    That is #358 §6, and it costs nothing: this reads the same grid `cell_text` already has.

    **The header row and column are a GUESS**, and the caller is told so in a caveat rather
    than left to assume. Row 1 and column A are the convention, and a sheet with a title block
    above the table, a transposed layout, or merged headers breaks it. Reporting a guess as a
    fact is the failure this project avoids elsewhere by saying `tab_ambiguous` instead of
    picking the first sheet.

    A header equal to the cell itself is suppressed: for a comment ON row 1 or column A the
    "header" would be the cell's own text, which reads as a fact and is noise.

    Indices are 1-BASED, like `_at` and like `Location.row`/`col`. Written 0-based first, which
    silently returned no headers at all rather than failing - the tests caught it.
    """
    row_header = _at(grid, row, 1) if col != 1 else ""
    column_header = _at(grid, 1, col) if row != 1 else ""
    return (row_header or None), (column_header or None)


def _at(grid: list[list[Any]], row: int, col: int) -> str:
    """The text at 1-based (row, col), or "" — which is what an empty cell holds.

    Empty rather than None on purpose: a cell inside the sheet with nothing in it *is* empty,
    and a register that prints "None" there is reporting a different fact.
    """
    if row < 1 or col < 1 or row > len(grid):
        return ""
    line = grid[row - 1] or []
    return "" if col > len(line) else str(line[col - 1])


def comment_rows(document: Any, comments: list,
                 contexts: list | None = None) -> tuple[list[str], list[dict], list[str]]:
    """(columns, rows, caveats)."""
    grids, tabs = _cell_lookup(document)
    caveats: list[str] = []

    rows: list[dict] = []
    anchored = 0                # comments that name a cell
    unplaced = 0                # ...of which, the ones whose tab could not be resolved
    for i, comment in enumerate(comments):
        location = getattr(comment, "location", None)
        cell = getattr(location, "cell", None) if location else None
        tab = getattr(location, "tab", None) if location else None
        cell_text = None
        by_tab = None
        row_header = None
        column_header = None
        if cell and grids:
            anchored += 1
            row_i = getattr(location, "row", 0)
            col_i = getattr(location, "col", 0)
            if tab in grids:
                # The tab is known (#290), so read only that tab. Offering every tab's
                # candidate alongside a settled answer invites the reader to second-guess it.
                cell_text = _at(grids[tab], row_i, col_i)
                row_header, column_header = _headers(grids[tab], row_i, col_i)
            elif len(tabs) == 1:
                cell_text = _at(grids[tabs[0]], row_i, col_i)
                row_header, column_header = _headers(grids[tabs[0]], row_i, col_i)
            else:
                # Unknown tab on a multi-tab workbook: the pre-#290 fallback. `tab` stays None
                # rather than guessing, and the content of each candidate usually settles it.
                tab = None
                unplaced += 1
                by_tab = {name: _at(grid, row_i, col_i) for name, grid in grids.items()}

        author = getattr(comment.author, "display_name", None) if comment.author else None
        rows.append(_row(tab=tab, thread_id=comment.id, reply_to=None, author=author,
                         created_time=_iso(getattr(comment, "created_time", None)),
                         resolved=bool(comment.resolved), text=comment.content,
                         quoted_text=getattr(comment, "quoted_text", None),
                         # TRUE/FALSE as text, never blank: a register cell has no third
                         # state, and this one is genuinely known for every comment.
                         anchored="TRUE" if getattr(comment, "anchored", False) else "FALSE",
                         anchor_state=getattr(comment, "anchor_state", ""),
                         cell=cell, cell_text=cell_text, cell_text_by_tab=by_tab,
                         row_header=row_header, column_header=column_header,
                         **_context_cells(contexts, i)))
        for reply in (comment.replies or []):
            reply_author = (getattr(reply.author, "display_name", None)
                            if getattr(reply, "author", None) else None)
            # A reply carries no passage and no cell of its own: only the top-level comment
            # anchors, and repeating the thread's anchor on every reply would make a register
            # look like several separate findings.
            rows.append(_row(thread_id=reply.id, reply_to=comment.id, author=reply_author,
                             created_time=_iso(getattr(reply, "created_time", None)),
                             resolved=bool(comment.resolved), text=reply.content))

    # Emitted AFTER the loop because it counts. The pre-#290 caveat fired on every multi-tab
    # workbook and said the file had no tab column at all; now most rows carry one, so a blanket
    # warning would cry wolf on exactly the files this fixed. It reports the shortfall instead.
    # TWO distinct failure modes, and they need different sentences. This one is "no cell at
    # all": the XLSX export the mapping depends on was unavailable, or every comment is
    # file-level. The old blanket multi-tab caveat covered this case by accident - it fired
    # whenever there was more than one tab - so narrowing that one without adding this would
    # have left a register with three empty columns and no explanation.
    if tabs and comments and not anchored:
        caveats.append(
            "No comment here could be mapped to a cell, so `cell`, `tab` and `cell_text` are "
            "empty on every row. Either these are file-level comments, which have no anchor at "
            "all, or the XLSX export that cell mapping depends on was unavailable for this "
            "file. The comments themselves are complete either way.")
    # THE SILENT-ZERO GUARD, and the reason it is a caveat rather than a column: notes are a
    # different object with no place in a reply-and-resolve register, so they do not belong in
    # the rows - but a register that omits them without saying so reports a sheet covered in
    # notes as having nothing to look at. #358 names this as the expensive failure, from
    # experience: a `resolved` field parsed against the wrong vocabulary once turned 17 closed
    # threads into 0 while every test stayed green.
    notes = _note_count(document)
    if notes:
        caveats.append(
            f"This spreadsheet also has {notes} cell NOTE(s), which are NOT in this register. A "
            f"note is a different kind of annotation: no author, no thread, and it cannot be "
            f"replied to or resolved - so nothing in a comment workflow applies to it. Use "
            f"`list_notes` to read them.")
    if any(r.get("row_header") or r.get("column_header") for r in rows):
        caveats.append(
            "`row_header` and `column_header` are a GUESS: they read column A and row 1, which "
            "is the usual layout and not a guaranteed one. A sheet with a title block above the "
            "table, a transposed layout, or merged headers will label a cell wrongly. Check "
            "them against `cell_text` before quoting them as the meaning of a cell.")
    if len(tabs) > 1 and unplaced:
        caveats.append(
            f"{unplaced} of {anchored} anchored comment(s) could not be placed on a tab: this "
            f"workbook has {len(tabs)} tabs ({', '.join(tabs)}) and the export did not say "
            f"which one those comments are on, so their `tab` is empty. `cell_text_by_tab` "
            f"shows what that cell holds on each tab instead - the content usually makes it "
            f"obvious. The other rows' `tab` is exact.")
    return list(COLUMNS), rows, caveats


def _iso(value: Any) -> str | None:
    return value.isoformat() if value is not None and hasattr(value, "isoformat") else None


# ── Destinations ────────────────────────────────────────────────────────────────────────
#
# "Here are some rows" saves nobody an hour. The two things a person actually wants are a CSV
# they can open and a Sheet they can share, so both are first-class rather than left to the
# caller to assemble.

CSV_SUFFIX = ".csv"


def flatten(value: Any) -> str:
    """One CSV cell's worth of text.

    `cell_text_by_tab` is a dict, and a CSV cell reading `{'Summary': '42'}` is not something a
    person can read - so it renders as `Summary=42 | Detail=Q3 revenue`, which is legible in a
    spreadsheet column and still says which tab held what.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, dict):
        return " | ".join(f"{k}={v}" for k, v in value.items())
    return str(value)


def to_csv(columns: list[str], rows: list[dict]) -> str:
    """RFC 4180 via the stdlib, so quoting and embedded newlines are somebody else's problem.

    `\r\n` because that is what `csv` writes by default and what Excel expects; Sheets and
    every other reader cope with it.
    """
    import csv as _csv
    import io
    buf = io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([csv_safe(flatten(row.get(c))) for c in columns])
    return buf.getvalue()


def to_grid(columns: list[str], rows: list[dict]) -> list[list[str]]:
    """Header row plus data, for a Sheets `values.update`."""
    return [list(columns)] + [[flatten(row.get(c)) for c in columns] for row in rows]


# ── Formula injection ───────────────────────────────────────────────────────────────────
#
# A comment on a shared document can begin with `=`, `+`, `-` or `@`, and Excel reads such a
# cell as a FORMULA when the file is opened - `=cmd|' /C calc'!A0` being the classic DDE
# payload. Anyone who can comment on a document we share can plant it, and the whole point of
# this feature is that a human opens the result in a spreadsheet. So this is not a theoretical
# concern here; it is the primary risk in SECURITY.md arriving by a new route.
#
# v0.24.0 shipped the CSV without this escape. The remedy is OWASP's: a leading apostrophe,
# which Excel and Sheets both read as "the rest is text" while leaving the value legible.
#
# NOT applied to `to_grid`, and the reason is narrower than it used to say here. This comment
# claimed "a Sheets write uses RAW" as a general fact. It was only ever LOCALLY true: it held
# for the one path `to_grid` uses - `Sheet.update`, whose default is RAW - while the MCP tool
# `update_cells` sitting above it defaulted to USER_ENTERED and would have evaluated the same
# text as a formula (#181, fixed by making RAW the default there too).
#
# So the premise is now true everywhere AND enforced, rather than asserted: the RAW-ness of
# this path is asserted by tests/test_raw_is_the_default.py, and the eight library declarations
# it depends on are asserted there too. Escaping here would put a stray apostrophe into
# somebody's spreadsheet, which is why the guard belongs on the write and not on the value.
#
# Note the escape sets are DELIBERATELY different per format and must stay that way: Excel
# reading a CSV acts on = + - @, openpyxl infers a formula from `=` alone, and a RAW Sheets
# write needs no escaping at all. One shared helper would be wrong in two directions.
FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


def csv_safe(text: str) -> str:
    return "'" + text if text[:1] in FORMULA_LEAD else text


def _slug(name: str) -> str:
    """A document title reduced to something usable as a filename.

    Titles are untrusted - somebody can name a Doc `../../etc/passwd` - so this keeps only
    characters that cannot mean anything to a path, and never returns empty.
    """
    keep = [c if (c.isalnum() or c in " -_") else " " for c in (name or "")]
    slug = " ".join("".join(keep).split())[:60].strip(" .-_")
    return slug or "comments"


def resolve_export_path(path: str | None, *, default_dir: str | None, doc_name: str,
                        stamp: str, suffix: str = CSV_SUFFIX):
    """(target, a sentence saying what happened) — or `ValueError` saying why not.

    A full path is ALLOWED, deliberately. A Claude Desktop *project* may only be able to write
    inside its own folder, where `~/Downloads` is unreachable; a Claude Code user wants the
    register in the repo they are working in. Confining to one directory would break exactly
    the cases where the file is most useful.

    What makes that safe is not validating the path but making every failure mode inert:

      * **nothing is ever overwritten** - an existing target gets `-<stamp>` appended, so the
        worst case is an unexpected file rather than a destroyed one;
      * **the extension is forced to `.csv`** - `~/.zshrc` becomes `~/.zshrc.csv`, which no
        shell will read;
      * **directories are never created**, so a path cannot conjure a tree;
      * the caller reports the resolved absolute path, because visibility is the last control.
    """
    from pathlib import Path
    given = (path or "").strip()
    told_where = False

    if not given:
        given = f"{_slug(doc_name)} comments {stamp}{suffix}"
        told_where = True
    has_dir = any(sep in given for sep in ("/", "\\")) or given.startswith("~")
    if has_dir:
        target = Path(given).expanduser()
    else:
        if not default_dir:
            raise ValueError("no default export directory is available; pass a full path")
        target = Path(default_dir).expanduser() / given
        told_where = True

    # `.csv` whatever was asked for. A name with no suffix gets one; a name with the wrong
    # suffix keeps it and gains ours, so `.zshrc` -> `.zshrc.csv` rather than `.csv`, which
    # would silently rename somebody's intended file.
    if target.suffix.lower() != suffix:
        target = target.with_name(target.name + suffix)

    parent = target.parent.expanduser()
    if not parent.is_dir():
        raise ValueError(
            f"{parent} does not exist or is not a directory, and directories are not created "
            f"automatically. Pass a full path to somewhere that exists, or use "
            f'destination="csv" to get the text back instead.')
    target = (parent.resolve() / target.name)

    note = ""
    if target.exists():
        target = target.with_name(f"{target.stem}-{stamp}{suffix}")
        note = (f"A file of that name already existed, so this was written as "
                f"{target.name} instead - nothing was overwritten. ")
    if told_where:
        note += f"Written to {target}. "
    return target, note


# ── Excel ───────────────────────────────────────────────────────────────────────────────
#
# A register somebody works through, rather than a CSV they have to import first. Asked for
# one, the honest answer used to be "CSV, a Google Sheet, or rows", and converting by hand hit
# two things the tool should handle - so it does.

XLSX_SUFFIX = ".xlsx"
HEADER_FILL = "1F3864"
# The cells somebody is meant to WRITE in. A register is mostly a read-only record, so the
# editable columns should not look like the rest of it - the same convention a financial model
# uses to mark its inputs.
INPUT_FILL = "FFF2CC"

# What each decision column offers. NOT symmetrical, deliberately:
#
#   resolve_comment  TRUE / FALSE / blank - three real states. TRUE resolves, FALSE REOPENS,
#                    and blank means "I have not decided", which most rows will be.
#   delete_comment   TRUE / blank only. Drive has no undelete for a comment, so offering FALSE
#                    would imply a reversal that does not exist - on the one action here that
#                    genuinely cannot be undone.
#
# Every value offered must be one `_apply.truthy` accepts, or the sheet hands somebody a value
# that then fails on import, which is worse than no dropdown at all.
# NO_CHANGE first, because it is what every cell already says and the list should open on it.
#
# `delete_comment` deliberately does NOT offer FALSE. Next to the word "delete", FALSE reads as
# *undo the delete* - and there is no undelete for a Drive comment, so it would invite that
# misreading on the one action here that genuinely cannot be reversed. A row that still says
# NO_CHANGE after somebody has been through it already records "considered, keeping it"; a
# second word for the same outcome would only add the wrong implication.
DROPDOWNS = {"resolve_comment": '"TRUE,FALSE,NO_CHANGE"',
             "delete_comment": '"TRUE,NO_CHANGE"'}

# Left BLANK on export, with the vocabulary in the dropdown instead. Blank has always been the
# safe default - it means no change - and pre-filling was considered for making that explicit,
# then dropped: `NO_CHANGE` on all 205 rows is noise you have to read past, and a column of
# blanks lets somebody see at a glance which rows they have touched. The dropdown carries the
# discoverability that a pre-filled value would have carried.
#
# (Pre-filling `FALSE` was never an option: on `resolve_comment` it means *reopen*, so a
# register filled with FALSE and applied untouched reopened every already-resolved thread.)
#
# `NO_CHANGE` remains selectable, for recording a decision rather than leaving a cell that
# cannot be told apart from one nobody reached.

# openpyxl refuses to write these and raises IllegalCharacterError. Reviewer text is arbitrary
# human input - pasted from terminals, editors, mail clients - so it genuinely contains them: a
# real 205-comment review had one. A register that will not be written because somebody pasted
# from a terminal is no register at all.
_ILLEGAL = "".join(chr(c) for c in list(range(0, 9)) + [11, 12] + list(range(14, 32)))
_STRIP = str.maketrans("", "", _ILLEGAL)


def _sheet_safe(value: Any) -> Any:
    text = flatten(value)
    return text.translate(_STRIP) if text else text


def used_columns(columns: list[str], rows: list[dict]) -> list[str]:
    """Only the columns that carry something.

    On a DOCUMENT the three Sheets-only columns (`cell`, `cell_text`, `cell_text_by_tab`) are
    structurally absent, and three empty columns on a review register suggest the export failed
    to fill them rather than that they do not apply. Computed from the data rather than from
    the file type, so it stays right for a type that has neither.
    """
    if not rows:
        return list(columns)
    # The action columns are kept whatever they contain: they are ALWAYS empty on export,
    # because they are what somebody fills in. Trimming them would remove the only part of the
    # sheet that is meant to be written to.
    keep = set(ACTIONS) | set(COMPLETED)
    return [c for c in columns
            if c in keep or any(flatten(row.get(c)) for row in rows)]


def _build_xlsx(columns: list[str], rows: list[dict], *, title: str):
    """Write a formatted, immediately usable register.

    **No formulas, and that is now enforced rather than merely intended.** Two separate
    reasons, and the second was missing until #182:

    *The register has no computed columns of its own.* openpyxl writes formulas with no cached
    values, so anything reading cached values - a thumbnail previewer, pandas - sees blanks
    until Excel opens the file and recalculates. A faithful register needs none, so it has
    none. Somebody wanting a pivot can build one on top.

    *And openpyxl INFERS cell type from value*, which the sentence above never contemplated: a
    comment body beginning `=` was written as a live formula element without anybody asking for
    one. Every data cell is therefore typed as text explicitly - see the comment on the write
    loop. The value is preserved byte for byte; the CSV path's apostrophe remedy would mangle
    the record to defend a reader that this format does not have.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:                      # pragma: no cover - environment-dependent
        raise ValueError(
            "writing .xlsx needs openpyxl, which is not installed. Either "
            "`pip install 'csa-google-workspace[xlsx]'`, or use destination=\"file\" for a "
            "CSV, which needs nothing extra.") from e

    keep = used_columns(columns, rows)
    wb = Workbook()
    ws = wb.active
    # A tab name may not exceed 31 characters or contain []:*?/\ - and it is derived from a
    # document title, which is untrusted.
    ws.title = ("".join(c for c in title if c not in "[]:*?/\\")[:31] or "Comments")
    ws.append(keep)
    for row in rows:
        ws.append([_sheet_safe(row.get(c)) for c in keep])
        # FORCE TEXT TYPING on every data cell. openpyxl INFERS type from value, so a comment
        # body beginning `=` was written as `<f>...</f>` - a live formula element - and a
        # spreadsheet opening the register evaluated it. Verified at the XLSX XML level rather
        # than by openpyxl's own re-read, since "reads back as a string" is a different claim
        # from "Excel treats it as text":
        #
        #   inferred          <c r="A2"><f>IMPORTXML(...)</f><v /></c>
        #   data_type='s'     <c r="A2" t="inlineStr"><is><t>=IMPORTXML(...)</t></is></c>
        #
        # Not an apostrophe prefix, unlike the CSV path: forcing the type leaves the value
        # byte-identical, and a register that mangles what somebody wrote is wrong about the
        # record. `set_explicit_value` does not exist in openpyxl 3.1.5 - assignment followed
        # by `data_type` is the mechanism that works there. (#182)
        for cell in ws[ws.max_row]:
            if isinstance(cell.value, str):
                cell.data_type = "s"

    font = "Arial"
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    for cell in ws[1]:
        cell.font = Font(name=font, bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # Frozen header and an autofilter, because two hundred unsorted rows is not a register.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 28

    # Dropdowns on the decision columns, so a value the importer would refuse cannot be typed.
    from openpyxl.worksheet.datavalidation import DataValidation
    last = ws.max_row
    for name, allowed in DROPDOWNS.items():
        if name not in keep or last < 2:
            continue
        letter = ws.cell(row=1, column=keep.index(name) + 1).column_letter
        # `showDropDown` is INVERTED against its name: the XML attribute means "suppress the
        # in-cell dropdown", so True HIDES the arrow. False is what shows it. Costly to
        # rediscover, because the workbook opens fine either way and only the arrow is missing.
        validation = DataValidation(type="list", formula1=allowed, allow_blank=True,
                                    showDropDown=False)
        validation.error = (f"{name} takes {allowed.strip(chr(34))} or nothing. A value nobody "
                            f"can read is refused on import rather than guessed at.")
        validation.errorTitle = "Not a value this column takes"
        ws.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}{last}")

    wide = {"text", "quoted_text", "cell_text_by_tab"}
    editable = set(ACTIONS)
    for i, name in enumerate(keep, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = 62 if name in wide else (
            22 if name in {"author", "created_time"} else 16)
    input_fill = PatternFill("solid", fgColor=INPUT_FILL)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            name = keep[cell.column - 1]
            cell.font = Font(name=font)
            cell.alignment = Alignment(vertical="top", wrap_text=name in wide)
            if name in editable:
                cell.fill = input_fill
    return wb


def xlsx_supported() -> bool:
    """Whether the formatted-workbook path is available (openpyxl installed).

    Asked rather than inferred from a caught exception, so a caller can choose a route *before*
    building anything. `openpyxl` ships with the `mcp` extra, so the server always has it; a
    library-only embedder may not, and a tool that used to work should degrade rather than fail.
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def to_xlsx(columns: list[str], rows: list[dict], target, *, title: str) -> None:
    """Write the formatted register to `target` (a path or a file object)."""
    _build_xlsx(columns, rows, title=title).save(target)


def to_xlsx_bytes(columns: list[str], rows: list[dict], *, title: str) -> bytes:
    """The same formatted register, as XLSX bytes, for upload-and-convert.

    Exists so `export_comments(destination="sheet")` can deliver the register Drive users get
    locally - header fill, frozen pane, autofilter, column widths, decision dropdowns - instead
    of the plain grid the values API can write. Sheets accepts neither Markdown nor styled text,
    so converting an uploaded workbook is the only route to a formatted spreadsheet.

    **Only ever for a file being CREATED.** Uploading an XLSX over an EXISTING spreadsheet
    silently resets every comment's cell anchor to A1 - measured 2026-08-31, see
    experiments/markdown-formatting/RESULTS.md. A brand-new file has no comments to lose, which
    is the only reason this route is safe here.
    """
    import io
    buf = io.BytesIO()
    _build_xlsx(columns, rows, title=title).save(buf)
    return buf.getvalue()
