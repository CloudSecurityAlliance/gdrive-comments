"""Apply a filled-in comment register back to the document.

The export makes 205 threads readable; this makes them actionable. It is also a **bulk mutation
of a shared document driven by a file**, which is a far sharper thing than an export, so most of
what follows is about the ways that goes wrong.

**Two layers of idempotency, because one is not enough.**

The obvious protection is a `*_completed` column ticked as each row is applied, so a re-run
skips finished work. That covers the ordinary case and fails in precisely the interesting one:
the reply posts, and the process dies *before* the tick is written. The sheet then says
not-done while the document says done, and a re-run trusting the marker alone posts the reply
twice — to a thread forty-two people are reading, with no way to unsend it.

So the marker is the **fast path** and the live document is the **authority**. Before posting,
look for a reply carrying this exact text, from this user, already on the thread. There is no
real reason to post a completely identical reply twice, so an exact match is treated as
evidence the work was already done — and `force` exists for whoever genuinely means it.

Resolve needs none of that: `resolved` *is* the state, so an already-resolved thread is skipped
on its own evidence.

**Everything is dry-run unless `apply` is set.** The blast radius is somebody's review, under
their name, and a spreadsheet is easy to get subtly wrong — a sort that did not carry every
column, a fill-down that overshot.
"""
from __future__ import annotations

import contextlib
import csv
import os
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from ._export import ACTIONS, COLUMNS, COMPLETED, XLSX_SUFFIX

DONE = "yes"

# The decision columns are THREE-state, and saying so in the type is the point. A bool that
# also had to carry "not decided" was how pre-filling `FALSE` came to mean "reopen every
# resolved thread" - an absence and a reversal are different things and one bool cannot hold
# both.
ACT = "act"          # TRUE      resolve · delete
REVERSE = "reverse"  # FALSE     reopen · (delete has no reversal - see decision())
NONE = "none"        # NO_CHANGE, or blank

# Pre-filled into the two decision columns on export. Safe to apply untouched, and unlike a
# blank cell it says what it means to somebody who has never seen the register before.
NO_CHANGE = "NO_CHANGE"

# Deliberately closed sets. "maybe later" in a resolve column must FAIL rather than be guessed
# at: guessing wrong closes somebody's open question.
TRUE = {"true", "yes", "y", "1", "x", "done", "✓"}
FALSE = {"false", "no", "n", "0"}
UNSET = {"", "no_change", "no change", "nochange", "none", "-", "n/a"}


def decision(value: Any) -> str | None:
    """ACT / REVERSE / NONE, or None when the cell says something nobody can read.

    BLANK counts as NONE alongside the sentinel, and must keep doing so: cells get cleared,
    sorts leave gaps, and a CSV can pass through a tool that drops a lone token. A default
    somebody can accidentally destroy is not a default.
    """
    text = _norm(value).lower()
    if text in UNSET:
        return NONE
    if text in TRUE:
        return ACT
    if text in FALSE:
        return REVERSE
    return None


@dataclass
class RowResult:
    # The SPREADSHEET row, header included - what a person navigates by. A 205-row register is
    # not something you scan, so "resolve_comment is 'maybe later'" without a row number leaves
    # somebody hunting for the cell that says it.
    row: int
    thread_id: str
    replied: bool = False
    resolved: bool = False
    reopened: bool = False
    deleted: bool = False
    failed: bool = False
    detail: str = ""


@dataclass
class Report:
    rows: list[RowResult] = field(default_factory=list)
    applied: bool = False
    wrong_file: str = ""

    def count(self, attr: str) -> int:
        return sum(1 for r in self.rows if getattr(r, attr))


def _norm(text: Any) -> str:
    """Whitespace-insensitive, because a spreadsheet cell round-trips with stray space.

    **Never `str(text or "")`.** `openpyxl` hands back a TRUE/FALSE cell as Python
    `True`/`False` - which is exactly what the xlsx register's own dropdown produces - and
    `False or ""` is `""`, so a boolean FALSE read as *blank* while a typed one read as
    REVERSE. Same intent, opposite behaviour, decided by a cell type nobody chose.

    Note the asymmetry that hid it for a release: `True or ""` is `True`, so `str()` gave
    `"True"` and the ACT path worked **by accident**. Only the reverse branch was broken.

    `0` and `0.0` had it too, and a numeric-formatted column is not exotic.
    """
    if text is None:
        return ""
    if isinstance(text, bool):          # BEFORE the numeric branch - bool subclasses int
        return "true" if text else "false"
    if isinstance(text, float) and text.is_integer():
        text = int(text)                # 1.0 -> "1", so a number cell reads like a typed one
    return " ".join(str(text).split())


def truthy(value: Any) -> bool | None:
    """True / False / None, where None means "not a value I will guess at"."""
    text = _norm(value).lower()
    if text in TRUE:
        return True
    if text in FALSE:
        return False
    return None


def read_rows(path: Path) -> list[dict]:
    """The register back, from .csv or .xlsx."""
    if path.suffix.lower() == XLSX_SUFFIX:
        try:
            from openpyxl import load_workbook
        except ImportError as e:                  # pragma: no cover - environment-dependent
            raise ValueError("reading .xlsx needs openpyxl: pip install "
                             "'csa-google-workspace[xlsx]'") from e
        ws = load_workbook(path, data_only=True).active
        it = ws.iter_rows(values_only=True)
        header = [str(h) if h is not None else "" for h in next(it, ())]
        # strict=False: a spreadsheet row may be short where trailing cells were never
        # touched, and a missing cell is an empty one, not a corrupt file.
        return [dict(zip(header, ["" if v is None else v for v in row], strict=False))
                for row in it]
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def write_back(path: Path, rows: list[dict], header: list[str]) -> None:
    """Rewrite the register with the completed markers, atomically.

    Temp file plus rename, because the whole point of the markers is surviving a crash, and a
    half-written register would be a worse state than the one being protected against.
    """
    if path.suffix.lower() == XLSX_SUFFIX:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        head = [c.value for c in next(ws.iter_rows(max_row=1))]
        index = {name: i for i, name in enumerate(head)}
        # strict=False deliberately: this runs AFTER the actions have been applied, so a
        # length mismatch must not raise and lose the report. The markers are the fast path
        # anyway - the document check is what actually prevents a double-post.
        for row_cells, row in zip(ws.iter_rows(min_row=2), rows, strict=False):
            for name in COMPLETED:
                if name in index and row.get(name):
                    row_cells[index[name]].value = row[name]
        _atomically(path, XLSX_SUFFIX, wb.save)
        return

    def write_csv(target: str) -> None:
        with open(target, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=header)
            writer.writeheader()
            for row in rows:
                writer.writerow({c: row.get(c, "") for c in header})

    _atomically(path, ".csv", write_csv)


def _atomically(path: Path, suffix: str, write: Any) -> None:
    """Write through a temp file in the same directory, then rename over the target.

    `mkstemp` plus an explicit `close`, NOT `NamedTemporaryFile`: on Windows, reopening a
    `NamedTemporaryFile` by name and replacing a file whose handle is still open BOTH raise
    `PermissionError` - and this runs *after* the document has been mutated, so the crash the
    completed markers exist to survive would be the very crash that stops them being written.
    Windows is supported here; the repo ships PowerShell installers. (#166)

    `delete=False` also orphaned the temp file whenever anything raised, leaving it next to
    somebody's register forever. Cleaned up on every path now.
    """
    fd, tmp = tempfile.mkstemp(dir=path.parent, suffix=suffix)
    os.close(fd)              # hand the file to the writer; Windows requires it closed first
    try:
        write(tmp)
        os.replace(tmp, path)          # no handle of ours is open on either side
    except BaseException:
        with contextlib.suppress(OSError):
            os.unlink(tmp)
        raise


def _mine_already_said(comment: Any, text: str) -> bool:
    """Has THIS user already posted this exact reply on this thread?

    Author-aware on purpose. The same text from somebody else is not evidence that I did the
    work - two reviewers can legitimately write "Fixed." - whereas my own identical reply
    almost certainly means the previous run got there and died before ticking the box.
    """
    wanted = _norm(text)
    for reply in (comment.replies or []):
        author = getattr(reply, "author", None)
        if getattr(author, "is_me", False) and _norm(reply.content) == wanted:
            return True
    return False


def _find_reply(comment: Any, reply_id: str) -> Any:
    """The reply object on a thread, or None. Replies obtained through a `Workspace` carry the
    backend, so `.delete()` works on them; a hand-built one would raise `DetachedError`."""
    for reply in (getattr(comment, "replies", None) or []):
        if reply.id == reply_id:
            return reply
    return None


def apply_rows(document: Any, rows: list[dict], *, apply: bool, force: bool) -> Report:
    """Walk the register once. Never raises for one bad row — 205 rows and #113 failing must
    not cost the other 204."""
    report = Report(applied=apply)
    by_id = {c.id: c for c in document.comments.all()}
    missing: list[int] = []

    for number, row in enumerate(rows, start=2):     # row 1 is the header
        thread = _norm(row.get("thread_id"))
        result = RowResult(row=number, thread_id=thread)
        reply_text = str(row.get("reply_comment") or "").strip()
        resolve_raw = row.get("resolve_comment")
        delete_raw = row.get("delete_comment")
        resolve = decision(resolve_raw)          # ACT | REVERSE | NONE | None
        delete = decision(delete_raw)
        wants_delete = delete is ACT

        notes: list[str] = []
        try:
            if _norm(row.get("reply_to")):
                parent_id = _norm(row.get("reply_to"))
                # Drive replies are flat - you reply to a THREAD, never to a reply - and
                # resolving acts on the thread. But DELETING a reply is a real operation, and
                # the one this column exists for: spam on a shared document usually arrives AS
                # a reply. Refusing it here used to send people to the parent row, which
                # deletes the WHOLE thread and strips every reply's text and author with it -
                # so following the advice destroyed other reviewers' work. (#170)
                if reply_text or resolve is not NONE:
                    raise ValueError(
                        f"this row is a reply. reply_comment and resolve_comment belong on the "
                        f"thread's own row, thread_id {parent_id!r} - Drive has no "
                        f"reply-to-a-reply, and resolving acts on the whole thread. "
                        f"delete_comment DOES work on this row and removes just this reply. "
                        f"Never move a delete to the parent row: that deletes the entire "
                        f"thread, including other people's replies, and cannot be undone.")
                if delete is None:
                    raise ValueError(
                        f"delete_comment is {str(delete_raw)!r}. Use TRUE to delete this reply "
                        f"or {NO_CHANGE} to leave it alone.")
                if delete is REVERSE:
                    raise ValueError(
                        f"delete_comment is {str(delete_raw)!r}, and there is no way to undo "
                        f"deleting a reply - Drive strips its text and author permanently. Use "
                        f"{NO_CHANGE} to leave it alone; nothing here can bring one back.")
                if delete is not ACT:
                    notes.append("a reply row; nothing to do")
                elif truthy(row.get("delete_comment_completed")):
                    notes.append("delete already marked done" + (
                        " (force does not apply - a deleted comment cannot be deleted again)"
                        if force else ""))
                else:
                    reply = _find_reply(by_id.get(parent_id), thread)
                    if reply is None:
                        raise ValueError(
                            f"no reply {thread!r} on thread {parent_id!r} in this file. Either "
                            f"the register came from a different document, or the reply is "
                            f"already gone.")
                    if getattr(reply, "deleted", False):
                        notes.append("already deleted")
                        row["delete_comment_completed"] = DONE
                    elif apply:
                        reply.delete()
                        row["delete_comment_completed"] = DONE
                        result.deleted = True
                        notes.append("reply deleted")
                    else:
                        result.deleted = True
                        notes.append("would delete this reply")
            elif not reply_text and resolve is NONE and delete is NONE:
                notes.append("no change requested")
            else:
                # The row's own validity first: an unreadable resolve value is wrong whether
                # or not the thread exists, and "maybe later" must fail rather than be guessed
                # at - guessing wrong closes somebody's open question.
                if resolve is None:
                    raise ValueError(
                        f"resolve_comment is {str(resolve_raw)!r}. Use TRUE to resolve, FALSE "
                        f"to reopen, or {NO_CHANGE} to leave it alone - a value nobody can "
                        f"read is not something to guess at when the result closes somebody's "
                        f"open question.")
                if delete is None:
                    raise ValueError(
                        f"delete_comment is {str(delete_raw)!r}. Use TRUE to delete or "
                        f"{NO_CHANGE} to leave it alone.")
                if delete is REVERSE:
                    # Refused rather than quietly treated as no-op. FALSE here almost certainly
                    # means "undo this delete", and there is no undelete for a Drive comment -
                    # silently doing nothing would leave somebody believing it worked.
                    raise ValueError(
                        f"delete_comment is {str(delete_raw)!r}, and there is no way to undo "
                        f"deleting a comment - Drive strips its text and author permanently. "
                        f"Use {NO_CHANGE} to leave a comment alone; nothing here can bring one "
                        f"back.")
                if wants_delete and (reply_text or resolve is not NONE):
                    # Replying to something you are about to destroy is incoherent, and doing
                    # one of the two silently would be a guess about which was meant.
                    raise ValueError(
                        "delete_comment is set alongside another action on the same row. "
                        "Deleting strips the text and the author permanently, so put it on "
                        "its own row or drop the other column.")
                # Reaching here means there IS something to do, so a missing thread is an
                # error unconditionally. (The earlier version guarded this with an `and`,
                # which was true but too subtle for a reader or a type checker to follow.)
                comment = by_id.get(thread)
                if comment is None:
                    # A DELETED comment is absent from a normal listing, so a re-run of a
                    # delete row would otherwise report "no such comment" - which reads like
                    # the wrong sheet rather than work already done. Ask again, including the
                    # deleted, BEFORE counting it as missing: counting first meant an ordinary
                    # re-run of a delete register tripped the wrong-document warning on every
                    # row, and a safety message that cries wolf on the normal path spends the
                    # attention it will need later. (#164)
                    if wants_delete:
                        try:
                            tombstone = document.comments.get(thread, include_deleted=True)
                        except Exception:         # noqa: BLE001 - genuinely absent
                            tombstone = None
                        if tombstone is not None and getattr(tombstone, "deleted", False):
                            notes.append("already deleted")
                            row["delete_comment_completed"] = DONE
                            report.rows.append(_finish(result, notes))
                            continue
                    missing.append(number)
                    raise ValueError(
                        f"no comment {thread!r} on this file. Most often that means the "
                        f"register was exported from a different document, or a sort moved the "
                        f"thread_id out of line with the rest of the row - check the fileId "
                        f"matches the document this sheet came from. Nothing on this row was "
                        f"changed.")

                if wants_delete:
                    if truthy(row.get("delete_comment_completed")):
                        notes.append("delete already marked done" + (
                            " (force does not apply - a deleted reply cannot be deleted again)"
                            if force else ""))
                    elif getattr(comment, "deleted", False):
                        notes.append("already deleted")
                        row["delete_comment_completed"] = DONE
                    elif apply:
                        comment.delete()
                        row["delete_comment_completed"] = DONE
                        result.deleted = True
                        notes.append("deleted")
                    else:
                        result.deleted = True
                        notes.append("would delete")
                    report.rows.append(_finish(result, notes))
                    continue

                # Reply BEFORE resolve: resolving posts its own visible action-reply, so the
                # substantive one has to land first or the thread reads backwards.
                if reply_text:
                    # `force` checked FIRST. The marker used to short-circuit ahead of it, so
                    # force only ever overrode the live-document duplicate check - and the
                    # person who genuinely means "say it again" is by definition working from a
                    # register that has already been applied, which is when markers exist. So
                    # force was inert in its own use case, and silently. (#169)
                    if not force and truthy(row.get("reply_comment_completed")):
                        notes.append("reply already marked done")
                    elif not force and _mine_already_said(comment, reply_text):
                        # The crash case: posted, then died before the tick.
                        notes.append("an identical reply from you is already on this thread; "
                                     "skipped (pass force to post it anyway)")
                        row["reply_comment_completed"] = DONE
                    elif apply:
                        comment.reply(reply_text)
                        row["reply_comment_completed"] = DONE
                        result.replied = True
                        notes.append("replied")
                    else:
                        result.replied = True
                        notes.append("would reply")

                # TRUE resolves, FALSE reopens, blank leaves it alone. `resolved` is the
                # state itself in both directions, so neither needs a marker to be idempotent -
                # the marker only saves a round trip.
                if resolve is ACT:
                    if comment.resolved:
                        notes.append("already resolved")
                        row["resolve_comment_completed"] = DONE
                    elif truthy(row.get("resolve_comment_completed")):
                        # Silence was the actual complaint in #169: "already marked done" is
                        # indistinguishable from force having been honoured and found nothing
                        # to do. Resolve is idempotent and an already-resolved thread is
                        # skipped on `comment.resolved` anyway, so there is nothing for force
                        # to mean here - which is now said rather than implied.
                        notes.append("resolve already marked done" + (
                            " (force does not apply to resolve - the thread's own resolved "
                            "state is the authority)" if force else ""))
                    elif apply:
                        comment.resolve()
                        row["resolve_comment_completed"] = DONE
                        result.resolved = True
                        notes.append("resolved")
                    else:
                        result.resolved = True
                        notes.append("would resolve")
                elif resolve is REVERSE:
                    if not comment.resolved:
                        notes.append("already open")
                        row["resolve_comment_completed"] = DONE
                    elif apply:
                        comment.reopen()
                        row["resolve_comment_completed"] = DONE
                        result.reopened = True
                        notes.append("reopened")
                    else:
                        result.reopened = True
                        notes.append("would reopen")

        except Exception as error:                # noqa: BLE001 - reported, never fatal
            result.failed = True
            # The flags are NOT reset. Anything already True records work that LANDED in a
            # shared document, and the register has already recorded it - so claiming
            # otherwise makes the report disagree with the artifact, and the report is what a
            # human reads. The original reasoning was half right: a refusal does have to be
            # unambiguous about whether it happened. It was made unambiguous by being untrue.
            # "Reply posted; resolve failed - re-run to finish" is both. (#163)
            did = [name for name in ("replied", "resolved", "reopened", "deleted")
                   if getattr(result, name)]
            text = str(error)
            if did:
                text = text.rstrip(". ") + (
                    f". ALREADY APPLIED on this row: {', '.join(did)} - that work is done and "
                    f"recorded in the register. Re-run to finish what is left; work already "
                    f"done is skipped, not repeated.")
            elif "nothing" not in text.lower():
                text += " Nothing on this row was changed."
            notes.append(f"{type(error).__name__}: {text}")

        report.rows.append(_finish(result, notes))

    # Said ONCE, at the top. When most rows name a thread this file does not have, the single
    # useful fact is "wrong register", and 205 identical row messages bury it. One bad id among
    # many good rows is a typo, not the wrong file - saying otherwise sends somebody looking
    # for a problem they do not have.
    if missing and len(missing) >= max(3, len(report.rows) // 2):
        report.wrong_file = (
            f"{len(missing)} of {len(report.rows)} rows name a comment this file does not "
            f"have. This register was most likely exported from a DIFFERENT DOCUMENT - check "
            f"the fileId. Nothing was changed on those rows.")
    return report


def _finish(result: RowResult, notes: list[str]) -> RowResult:
    result.detail = "; ".join(notes) or "nothing to do"
    return result


def header_for(rows: list[dict]) -> list[str]:
    """The sheet's own column order, extended with anything we must write back."""
    seen = list(rows[0]) if rows else list(COLUMNS)
    for name in (*ACTIONS, *COMPLETED):
        if name not in seen:
            seen.append(name)
    return seen
